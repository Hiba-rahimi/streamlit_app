# librairies
import pandas as pd
from parser_TT140_MasterCard import *
import os
import re
import tempfile
from openpyxl.styles import PatternFill
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from openpyxl.worksheet.table import Table, TableStyleInfo
#import win32com.client as win32
import streamlit as st
import io

pd.set_option('future.no_silent_downcasting', True)
pd.options.display.float_format = '{:,.2f}'.format

# Define the function to read CSV files with delimiters
def read_csv_with_delimiters(file_path, default_columns=None, default_delimiter=','):
    """
    Read a CSV file with delimiters `;`, `,`, or space.
    
    Parameters:
        file_path (str): The path to the CSV file.
        default_delimiter (str): The default delimiter to use if the file is empty.
        
    Returns:
        pd.DataFrame: The DataFrame with the CSV content.
    """
    try:
        with open(file_path, 'r') as f:
            first_line = f.readline()

        # Detect delimiter
        if ';' in first_line:
            delimiter = ';'
        elif ',' in first_line:
            delimiter = ','
        elif ' ' in first_line:
            delimiter = r'\s+'  # regex for one or more spaces
        else:
            delimiter = default_delimiter
    except FileNotFoundError:
        delimiter = default_delimiter

    try:
        df = pd.read_csv(file_path, sep=delimiter, engine='python' , thousands = ',' , decimal= ".")
    except pd.errors.EmptyDataError:
        df = pd.DataFrame(columns=default_columns)
    
    return df


# Function to save uploaded file to a temporary location
def save_uploaded_file(uploaded_file):
    with tempfile.NamedTemporaryFile(delete=False) as temp_file:
        temp_file.write(uploaded_file.read())
        return temp_file.name

# standard columns for each source file
default_columns_cybersource = ['NBRE_TRANSACTION', 'MONTANT_TOTAL', 'CUR', 'FILIALE', 'RESEAU', 'TYPE_TRANSACTION']
default_columns_saisie_manuelle = ['NBRE_TRANSACTION', 'MONTANT_TOTAL', 'CUR', 'FILIALE', 'RESEAU']
default_columns_pos = ['FILIALE', 'RESEAU', 'TYPE_TRANSACTION', 'DATE_TRAI', 'CUR', 'NBRE_TRANSACTION', 'MONTANT_TOTAL']

def reading_cybersource(cybersource_file):
    if os.path.exists(cybersource_file):
        df_cybersource = read_csv_with_delimiters(cybersource_file, default_columns_cybersource)
        df_cybersource.columns = df_cybersource.columns.str.strip()
        df_cybersource['TYPE_TRANSACTION'] = 'ACHAT'
        df_cybersource = df_cybersource.apply(lambda x: x.str.strip() if x.dtype == "object" else x)
        # handle unified format of columns for merging after
        df_cybersource['RESEAU'] = df_cybersource['RESEAU'].astype(str)
        df_cybersource['FILIALE'] = df_cybersource['FILIALE'].astype(str)
        df_cybersource['CUR'] = df_cybersource['CUR'].astype(str)
        return df_cybersource
    else:
        df_cybersource = pd.DataFrame(columns=default_columns_saisie_manuelle)
        #st.write("The Saisie Manuelle file does not exist at the specified path.")

# Read Saisie Manuelle file
def reading_saisie_manuelle(saisie_manuelle_file):
    if os.path.exists(saisie_manuelle_file):
        df_sai_manuelle = read_csv_with_delimiters(saisie_manuelle_file, default_columns_saisie_manuelle)
        df_sai_manuelle.columns = df_sai_manuelle.columns.str.strip()
        df_sai_manuelle = df_sai_manuelle.apply(lambda x: x.str.strip() if x.dtype == "object" else x)
        # handle unified format of columns for merging after
        df_sai_manuelle['RESEAU'] = df_sai_manuelle['RESEAU'].astype(str)
        df_sai_manuelle['FILIALE'] = df_sai_manuelle['FILIALE'].astype(str)
        df_sai_manuelle['CUR'] = df_sai_manuelle['CUR'].astype(str)
        
        return df_sai_manuelle
    else:
        df_sai_manuelle = pd.DataFrame(columns=default_columns_saisie_manuelle)
        #print("The Saisie Manuelle file does not exist at the specified path.")

# Read POS file
def reading_pos(pos_file):
    if os.path.exists(pos_file):
        df_pos = read_csv_with_delimiters(pos_file, default_columns_pos)
        df_pos.columns = df_pos.columns.str.strip()
        df_pos = df_pos.apply(lambda x: x.str.strip() if x.dtype == "object" else x)
        df_pos.rename(columns={'BANQUE': 'FILIALE'}, inplace=True)
        # handle unified format of columns for merging after
        df_pos['RESEAU'] = df_pos['RESEAU'].astype(str)
        df_pos['FILIALE'] = df_pos['FILIALE'].astype(str)
        df_pos['CUR'] = df_pos['CUR'].astype(str)
        return df_pos
    else:
        df_pos = pd.DataFrame(columns=default_columns_pos)
        #print("The POS file does not exist at the specified path.")

def filtering_sources(df_cybersource, df_sai_manuelle, df_pos):
    # Filter each source to get MASTERCARD network transactions
    filtered_cybersource_df = df_cybersource[df_cybersource['RESEAU'] == 'MASTERCARD INTERNATIONAL']
    filtered_saisie_manuelle_df = df_sai_manuelle[df_sai_manuelle['RESEAU'] == 'MASTERCARD INTERNATIONAL']
    filtered_pos_df = df_pos[(df_pos['RESEAU'] == 'MASTERCARD INTERNATIONAL') & 
                         (~df_pos['TYPE_TRANSACTION'].str.endswith('_MDS'))]
    return filtered_cybersource_df, filtered_saisie_manuelle_df, filtered_pos_df


def validate_file_name_and_date(file_name, source, date_to_validate=None):
    """
    Validate the file name based on the source, the required pattern, 
    and optionally validate the date extracted from the file name.
    
    Parameters:
        file_name (str): The name of the uploaded file.
        source (str): The source type (CYBERSOURCE, POS, or SAIS_MANU).
        date_to_validate (str): The date to validate against the one in the file name. (Optional)
        
    Returns:
        bool: True if the file name is valid, False otherwise.
        
    Raises:
        ValueError: If the file name is invalid or if the date does not match the expected pattern.
    """
    pattern = f"^TRANSACTION_{source}_TRAITE_SG_\\d{{2}}-\\d{{2}}-\\d{{2}}_\\d{{6}}\\.CSV$"
    if not re.match(pattern, file_name):
        raise ValueError(f"Invalid file name: {file_name}. Expected pattern: TRANSACTION_{source}_TRAITE_SG_YY-MM-DD_HHMMSS.CSV")

    # Extract the date from the file name
    date_match = re.search(r"\d{2}-\d{2}-\d{2}", file_name)
    if date_match:
        extracted_date = date_match.group(0)
        if date_to_validate:
            if extracted_date != date_to_validate:
                raise ValueError(f"Date extracted from the file name ({extracted_date}) does not match the provided date ({date_to_validate}).")
    else:
        raise ValueError("Date not found in the file name.")

    return True

# Converting the excel rejects file to a df
def excel_to_csv_to_df(excel_file_path, sheet_name=0):
    """
    Converts an Excel file to a CSV file and then reads it into a Pandas DataFrame.

    Parameters:
        excel_file_path (str): Path to the Excel file.
        sheet_name (str or int): Name or index of the sheet to convert (default is the first sheet).

    Returns:
        pd.DataFrame: DataFrame containing the data from the CSV file.
    """
    try:
        # Read the Excel file
        df = pd.read_excel(excel_file_path, sheet_name=sheet_name, header=0)
        
        # Define the CSV file path
        csv_file_path = excel_file_path.replace('.xlsx', '.csv')
        
        # Save the DataFrame to a CSV file without an index
        df.to_csv(csv_file_path, index=False)
        
        # Read the CSV file into a DataFrame
        df_csv = read_csv_with_delimiters(csv_file_path)
        return df_csv
    
    except PermissionError as p_error:
        print(f"Permission error: {p_error}. Please check your file permissions.")
    except Exception as e:
        print(f"An error occurred: {e}")
    return None

def standardize_date_format(date_column, desired_format='%Y-%m-%d'):
    """
    Standardize the date format in a given column.
    
    Parameters:
        date_column (pd.Series): The column containing dates to standardize.
        desired_format (str): The desired date format (default is '%Y-%m-%d').
    
    Returns:
        pd.Series: The column with dates in the standardized format.
    """
    # Convert all dates to datetime objects
    date_column = pd.to_datetime(date_column , dayfirst=False  , yearfirst=True)
    # Format all datetime objects to the desired format
    date_column = date_column.dt.strftime(desired_format )
    
    return date_column


# Function to format columns
def format_columns(df):
    for col in df.columns:
        if 'Montant' in col:
            df[col] = df[col].apply(lambda x: '{:,.2f}'.format(x) if isinstance(x, (int, float)) else x)
    return df

# Merge the dataframes on relevant common columns
def merging_sources_without_recycled(filtered_cybersource_df, filtered_saisie_manuelle_df, filtered_pos_df):
# Merge POS and Saisie Manuelle dataframes
    merged_df = pd.merge(
        filtered_pos_df,
        filtered_saisie_manuelle_df,
        on=['FILIALE', 'RESEAU', 'CUR'],
        suffixes=('_pos', '_saisie'),
        how='outer'  # Use outer join to keep all rows from both dataframes
    )

    # Merge with filtered_cybersource_df
    merged_df = pd.merge(
        merged_df,
        filtered_cybersource_df,
        on=['FILIALE', 'RESEAU', 'CUR', 'TYPE_TRANSACTION'],
        suffixes=('_merged', '_cybersource'),
        how='outer'  # Use outer join to keep all rows from all dataframes
    )

    # Fill missing values with 0 and sum the 'NBRE_TRANSACTION' values
    merged_df['NBRE_TRANSACTION'] = (merged_df['NBRE_TRANSACTION_pos'].fillna(0) +
                                    merged_df['NBRE_TRANSACTION_saisie'].fillna(0) +
                                    merged_df['NBRE_TRANSACTION'].fillna(0))

    # Convert 'NBRE_TRANSACTION' to integer
    merged_df['NBRE_TRANSACTION'] = merged_df['NBRE_TRANSACTION'].astype(int)

    # Use MONTANT_TOTAL from filtered_pos_df
    merged_df['MONTANT_TOTAL'] = merged_df['MONTANT_TOTAL_pos'].fillna(0).infer_objects()


    # Use MONTANT_TOTAL from filtered_pos_df
    merged_df['MONTANT_TOTAL'] = merged_df['MONTANT_TOTAL_pos'].fillna(0)

    # Drop unnecessary columns
    merged_df.drop(['NBRE_TRANSACTION_pos', 'NBRE_TRANSACTION_saisie', 'MONTANT_TOTAL_pos', 'MONTANT_TOTAL_saisie'], axis=1, inplace=True)

    # Drop duplicate rows
    merged_df.drop_duplicates(subset=['FILIALE', 'RESEAU', 'CUR', 'TYPE_TRANSACTION', 'DATE_TRAI'], inplace=True)
    total_nbre_transactions = merged_df['NBRE_TRANSACTION'].sum()
    
    merged_df = merged_df.reset_index(drop=True)

    return merged_df , total_nbre_transactions


def merging_with_recycled(recycled_rejected_file,filtered_cybersource_df, filtered_saisie_manuelle_df, filtered_pos_df, filtering_date):
    df_merged, _ = merging_sources_without_recycled(filtered_cybersource_df, filtered_saisie_manuelle_df, filtered_pos_df)
    df_recycled = excel_to_csv_to_df(recycled_rejected_file)    
    df_recycled.columns = df_recycled.columns.str.strip()
    df_recycled = df_recycled.apply(lambda x: x.str.strip() if x.dtype == "object" else x)
    df_recycled.rename(columns={'BANQUE': 'FILIALE'}, inplace=True)
    df_recycled['Date Retraitement'] = standardize_date_format(df_recycled['Date Retraitement'])
    # Filter rows where 'Date_Retraitement' is not equal to the specified date
    
    df_recycled = df_recycled[df_recycled['Date Retraitement'] == filtering_date.strftime('%Y-%m-%d')]
    
    df_recycled.drop_duplicates(subset=['FILIALE', 'RESEAU', 'ARN', 'Autorisation', 'Date Transaction', 'Montant', 'Devise'], inplace=True)
    # Normalize 'FILIALE' values
    df_recycled['FILIALE'] = df_recycled['FILIALE'].str.replace("COTE D'IVOIRE", "COTE D IVOIRE")

    # Remove any commas and spaces from the 'Montant' column and convert it to numeric
    # df_recycled['Montant'] = df_recycled['Montant'].astype(str)  # Ensure all values are strings
    # df_recycled['Montant'] = df_recycled['Montant'].str.replace(',', '').str.replace(' ', '')  # Remove commas and spaces
    #df_recycled['Montant'] = df_recycled['Montant'].astype(float)  # Convert to float

    # Normalize 'FILIALE' values
    df_recycled['FILIALE'] = df_recycled['FILIALE'].str.replace('SG-', 'SG - ')

    # Group by FILIALE and RESEAU and calculate the count and sum of Montant
    summary = df_recycled.groupby(['FILIALE', 'RESEAU']).agg(
        NBRE_TRANSACTION=('Montant', 'count'),
        MONTANT_TOTAL=('Montant', 'sum')
    ).reset_index()

    # Merge the summary DataFrame into the corresponding 'FILIALE' values of the 'df_merged' DataFrame
    merged_df = df_merged.merge(summary, on=['FILIALE', 'RESEAU'], how='left', suffixes=('_merged', '_summary'))

    # Fill NaN values with 0
    merged_df.fillna(0, inplace=True)

    # Sum the values of 'NBRE_TRANSACTION_merged' and 'NBRE_TRANSACTION_summary' columns and assign it to 'NBRE_TRANSACTION'
    merged_df['NBRE_TRANSACTION'] = merged_df['NBRE_TRANSACTION_merged'] + merged_df['NBRE_TRANSACTION_summary']

    # Drop unnecessary columns
    merged_df.drop(['NBRE_TRANSACTION_merged', 'NBRE_TRANSACTION_summary'], axis=1, inplace=True)

    # Sum the values of 'MONTANT_TOTAL_merged' and 'MONTANT_TOTAL_summary' columns and assign it to 'MONTANT_TOTAL'
    merged_df['MONTANT_TOTAL'] = merged_df['MONTANT_TOTAL_merged'] + merged_df['MONTANT_TOTAL_summary']

    # Drop unnecessary columns
    merged_df.drop(['MONTANT_TOTAL_merged', 'MONTANT_TOTAL_summary'], axis=1, inplace=True)

    # Convert 'NBRE_TRANSACTION' to integers
    merged_df['NBRE_TRANSACTION'] = merged_df['NBRE_TRANSACTION'].astype(int)

    total_nbre_transactions = merged_df['NBRE_TRANSACTION'].sum()
    #print(total_nbre_transactions)


    return df_recycled, merged_df, total_nbre_transactions


def populating_table_reconcialited(merged_df):
    # Columns to be included in the reconciliated DataFrame
    new_columns = [
        'FILIALE', 'Réseau', 'Type', 'Date', 'Devise', 'Nbre Total De Transactions',
        'Montant Total de Transactions', 'Rapprochement', 'Nbre Total de Rejets',
        'Montant de Rejets', 'Nbre de Transactions (Couverture)', 
        'Montant de Transactions (Couverture)'
    ]

    # Mapping between the old and new column names
    column_mapping = {
        'FILIALE': 'FILIALE',
        'RESEAU': 'Réseau',
        'TYPE_TRANSACTION': 'Type',
        'DATE_TRAI': 'Date',
        'CUR': 'Devise',
        'NBRE_TRANSACTION': 'Nbre Total De Transactions',
        'MONTANT_TOTAL': 'Montant Total de Transactions'
    }

    # Rename columns
    merged_df.rename(columns=column_mapping, inplace=True)

    # Add the missing columns with default values
    for column in set(new_columns) - set(merged_df.columns):
        merged_df[column] = ''

    # Reorder the columns
    merged_df = merged_df[new_columns]
    return merged_df

def handle_exact_match_csv(merged_df , run_date):
    populating_table_reconcialited(merged_df)
    df_reconciliated = merged_df.copy()
    run_date_new = pd.to_datetime(run_date, format='%y-%m-%d')
    formatted_date = run_date_new.strftime('%Y-%m-%d')
    df_reconciliated['Date'] = formatted_date
    df_reconciliated['Rapprochement'] = 'ok'
    df_reconciliated['Montant de Transactions (Couverture)'] = df_reconciliated['Montant Total de Transactions']
    df_reconciliated['Nbre de Transactions (Couverture)'] = df_reconciliated['Nbre Total De Transactions']
    df_reconciliated = format_columns(df_reconciliated)
    #df_reconciliated.to_csv('reconciliated.csv', index=False)
    return df_reconciliated

def handle_non_match_reconciliation(file_path,merged_df , run_date):
    populating_table_reconcialited(merged_df)
    df_reconciliated = merged_df.copy()
    # Load the rejected summary data
    df_rejected_summary = calculate_rejected_summary(file_path)

    # Ensure the relevant columns exist in the reconciliated DataFrame
    if 'FILIALE' not in df_reconciliated.columns:
        raise KeyError('The required columns are missing in the reconciliated DataFrame.')

    # Add the columns if they do not exist
    if 'Rapprochement' not in df_reconciliated.columns:
        df_reconciliated['Rapprochement'] = 'ok'

    # Create a set of FILIALEs with issues from the rejected summary
    filiales_with_issues = set(df_rejected_summary['FILIALE'])

    # Update the reconciliated DataFrame
    df_reconciliated['Rapprochement'] = df_reconciliated['FILIALE'].apply(
        lambda x: 'NOT OK' if x in filiales_with_issues else 'OK'
    )

    # Update Nbre Total de Rejets and Montant de Rejets based on the rejected summary data
    for index, row in df_rejected_summary.iterrows():
        filiale = row['FILIALE']
        nbr_rejets = row['Nbre Total de Rejets']
        montant_rejets = row['Montant de Rejets']

        # Find the matching row in the reconciliated DataFrame
        match_idx = df_reconciliated[df_reconciliated['FILIALE'] == filiale].index

        if not match_idx.empty:
            # Update the matching row
            df_reconciliated.loc[match_idx, 'Nbre Total de Rejets'] = nbr_rejets
            df_reconciliated.loc[match_idx, 'Montant de Rejets'] = montant_rejets
            df_reconciliated['Montant de Transactions (Couverture)'] = df_reconciliated['Montant Total de Transactions']
            df_reconciliated['Nbre de Transactions (Couverture)'] = df_reconciliated['Nbre Total De Transactions']

    # Fill NaN values in 'Nbre Total de Rejets' with 0 before converting to integer type
    df_reconciliated['Nbre Total de Rejets'] = df_reconciliated['Nbre Total de Rejets'].replace('', 0).fillna(0).astype(int)
    run_date_new = pd.to_datetime(run_date, format='%y-%m-%d')
    formatted_date = run_date_new.strftime('%Y-%m-%d')
    df_reconciliated['Date'] = formatted_date
    df_reconciliated = format_columns(df_reconciliated)
    #Save the updated DataFrame to a CSV file
    #df_reconciliated.to_csv('reconciliated.csv', index=False)
    return df_reconciliated


def blue_style_and_save_to_excel(df):
    """
    Styles a DataFrame and saves it to an Excel file with a predefined style.

    Parameters:
    - df (pd.DataFrame): The DataFrame to be styled and saved.

    Returns:
    - str: Path to the saved Excel file.
    """
    # Define the path for the output Excel file
    excel_path = './styled_data.xlsx'

    # Save DataFrame to an Excel file
    df.to_excel(excel_path, index=False)

    # Load the workbook and select the active worksheet
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook.active

    # Define the table style
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=True
    )

    # Add the table to the worksheet
    tab = Table(displayName="Table1", ref=sheet.dimensions)
    tab.tableStyleInfo = style
    sheet.add_table(tab)
    # Define the number format for thousands and decimal separators
    number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

    # Apply additional styling
    # Set column widths based on content length and apply number format
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
            # Apply number format to columns containing 'Montant'
            if 'Montant' in df.columns[col[0].column - 1]:
                cell.number_format = number_format
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width
        
    # Set the header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Save the styled workbook
    workbook.save(excel_path)

    return excel_path

def styling_and_saving_reconciliated(excel_path):
    """
    Styles specific cells in an existing Excel file based on 'Rapprochement' column values.

    Parameters:
    - excel_path (str): Path to the Excel file to be styled.

    Returns:
    - str: Path to the styled Excel file.
    """
    # Reopen the workbook with openpyxl to apply styles
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook['Sheet1']

    # Load DataFrame from the Excel file
    df = pd.read_excel(excel_path, sheet_name='Sheet1')
    
    # Apply styles to specific cells
    for row_idx, row in df.iterrows():
        for col_idx in range(len(row)):
            cell = sheet.cell(row=row_idx + 2, column=col_idx + 1)  # +2 to account for header and 1-based index
            # Apply bold font to 'Rapprochement' column cells
            if col_idx == row.index.get_loc('Rapprochement'):
                cell.font = Font(bold=True)
            # Apply red background for 'not ok' and white text color
            if row['Rapprochement'] == 'NOT OK':
                cell.fill = PatternFill(start_color='ffe26b0a', end_color='ffe26b0a', fill_type="solid")  # Red
                cell.font = Font(bold=True ,color="FFFFFF")  # Set text color to white

    # Save the styled workbook
    workbook.save(excel_path)
    return excel_path

            
def highlight_non_reconciliated_row(row):
    return [f'background-color: #ffab77; font-weight: bold;'
        if row['Rapprochement'] == 'NOT OK' else '' for _ in row]


df_reconciliated = None

def download_file(recon  , df, file_partial_name, button_label , run_date):
    # Assuming styling_and_saving_reconciliated is a defined function that processes the DataFrame
    excel_path1 = blue_style_and_save_to_excel(df)
    
    if recon : 
        excel_path2 = styling_and_saving_reconciliated(excel_path1)
    else:
        excel_path2 = excel_path1
    
    with open(excel_path2, 'rb') as f:
        excel_data = io.BytesIO(f.read())

    
    with open(excel_path2, 'rb') as f:
        excel_data = io.BytesIO(f.read())

    # Define the file name
    file_name = f"{file_partial_name}_{run_date}.xlsx"

    # Create a download button for the Excel file
    st.download_button(
        label=button_label,
        data=excel_data,
        file_name=file_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True
    )
    return excel_path2 , file_name
import tempfile
def save_excel_locally(excel_path , file_name):
    wb = openpyxl.load_workbook(excel_path)


    
    # Save the workbook locally with original file name
    temp_dir = tempfile.gettempdir()  # Get the temporary directory
    file_path = os.path.join(temp_dir, file_name)  # Define the file path
    
    wb.save(file_path)  # Save the workbook
    
    return file_path  # Return the file path

# def send_excel_contents_to_outlook(excel_path , file_name):
#     try:
#         # Save Excel file locally
#         excel_file_path = save_excel_locally(excel_path , file_name)
#
#         # Connect to Outlook
#         outlook = win32.Dispatch("Outlook.Application")
#
#         # Create a new email
#         mail = outlook.CreateItem(0)
#
#         # Attach Excel file
#         mail.Attachments.Add(excel_file_path)
#
#         # Display the Outlook application with the composed email
#         mail.Display(True)  # True opens the email in a new window
#
#     except Exception as e:
#         st.error(f"Error occurred while sending the email: {e}")
    # def show_modal_confirmation(modal_key, title, message, confirm_action, data, insert_type):
    #     modal = Modal(key=modal_key, title=title)
    #
    # if st.button(f':floppy_disk: Stocker résultats de {insert_type}', key=f'sauvegarder_resultats_{modal_key}',type="primary", use_container_width=True):
    #     modal.open()
    #
    # if modal.is_open():
    #     with modal.container():
    #         st.write(message)
    #         col1, col2 = st.columns(2)
    #         with col1:
    #             if st.button("Yes", key=f"{modal_key}_yes"):
    #                 confirm_action(data)
    #                 modal.close()
    #
    #         with col2:
    #             if st.button("No", key=f"{modal_key}_no"):
    #                 modal.close()
    #
    # return modal